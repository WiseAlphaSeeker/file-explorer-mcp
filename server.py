"""
Serveur MCP de gestion de fichiers
Un exemple simple et fonctionnel de serveur MCP.

Ce serveur expose 7 tools et 1 resource :
  - list_files      : lister les fichiers du repertoire
  - read_file       : lire le contenu d un fichier
  - search_in_files : chercher un mot dans tous les fichiers
  - file_stats      : obtenir des statistiques sur un fichier
  - edit_file       : modifier le contenu d un fichier texte
  - read_csv        : lire et filtrer un fichier CSV
  - edit_xlsx       : modifier une cellule dans un fichier Excel

Resource :
  - file://{filename} : acceder au contenu d un fichier comme ressource
"""

import os
import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Optional, List
from openpyxl import load_workbook

from mcp.server.fastmcp import FastMCP
from pydantic import BaseModel, Field

# Configuration
FILES_DIR = Path(__file__).parent / "sample-files"

# Initialisation du serveur MCP
# IMPORTANT : variable renommee 'server' au lieu de 'mcp'
# pour eviter le conflit avec le module 'mcp'
server = FastMCP("files_mcp")


# Tool 1 : Lister les fichiers
@server.tool(name="list_files")
async def list_files() -> str:
    """Liste tous les fichiers disponibles dans le repertoire."""
    files = []
    for f in sorted(FILES_DIR.iterdir()):
        if f.is_file():
            stat = f.stat()
            files.append({
                "name": f.name,
                "size_bytes": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })

    if not files:
        return "Aucun fichier trouve."

    lines = [f"{len(files)} fichier(s) trouve(s) :\n"]
    for f in files:
        lines.append(f"  - {f['name']}  ({f['size_bytes']} octets, modifie le {f['modified']})")
    return "\n".join(lines)


# Tool 2 : Lire un fichier
class ReadFileInput(BaseModel):
    filename: str = Field(..., description="Nom du fichier a lire (ex: notes.txt)", min_length=1)


@server.tool(name="read_file")
async def read_file(params: ReadFileInput) -> str:
    """Lit et retourne le contenu complet d un fichier."""
    filepath = FILES_DIR / params.filename

    if ".." in params.filename or "/" in params.filename:
        return "Erreur : Le nom de fichier ne peut pas contenir .. ou /."

    if not filepath.exists():
        available = [f.name for f in FILES_DIR.iterdir() if f.is_file()]
        return f"Erreur : Fichier '{params.filename}' introuvable. Disponibles : {', '.join(available)}"

    content = filepath.read_text(encoding="utf-8")
    return f"--- {params.filename} ---\n\n{content}"


# Tool 3 : Chercher dans les fichiers
class SearchInput(BaseModel):
    query: str = Field(..., description="Mot ou phrase a chercher", min_length=1)
    case_sensitive: Optional[bool] = Field(default=False, description="Sensible a la casse")


@server.tool(name="search_in_files")
async def search_in_files(params: SearchInput) -> str:
    """Recherche un mot ou une phrase dans tous les fichiers."""
    results = []
    query = params.query if params.case_sensitive else params.query.lower()

    for filepath in sorted(FILES_DIR.iterdir()):
        if not filepath.is_file():
            continue
        try:
            content = filepath.read_text(encoding="utf-8")
            for i, line in enumerate(content.splitlines(), 1):
                compare_line = line if params.case_sensitive else line.lower()
                if query in compare_line:
                    results.append({
                        "file": filepath.name,
                        "line_number": i,
                        "content": line.strip(),
                    })
        except UnicodeDecodeError:
            continue

    if not results:
        return f"Aucun resultat pour '{params.query}'."

    lines = [f"{len(results)} resultat(s) pour '{params.query}' :\n"]
    for r in results:
        lines.append(f"  - {r['file']} (ligne {r['line_number']}): {r['content']}")
    return "\n".join(lines)


# Tool 4 : Statistiques d un fichier
class FileStatsInput(BaseModel):
    filename: str = Field(..., description="Nom du fichier a analyser", min_length=1)


@server.tool(name="file_stats")
async def file_stats(params: FileStatsInput) -> str:
    """Retourne des statistiques detaillees sur un fichier."""
    filepath = FILES_DIR / params.filename
    if not filepath.exists():
        return f"Erreur : Fichier '{params.filename}' introuvable."

    content = filepath.read_text(encoding="utf-8")
    lines = content.splitlines()
    words = content.split()
    stat = filepath.stat()

    return json.dumps({
        "filename": params.filename,
        "extension": filepath.suffix,
        "size_bytes": stat.st_size,
        "lines": len(lines),
        "words": len(words),
        "characters": len(content),
        "empty_lines": sum(1 for l in lines if not l.strip()),
        "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
    }, indent=2, ensure_ascii=False)


# Tool 5 : Modifier un fichier
class EditFileInput(BaseModel):
    filename: str = Field(..., description="Nom du fichier a modifier (ex: notes.txt)", min_length=1)
    old_text: str = Field(..., description="Texte existant a remplacer (doit correspondre exactement)", min_length=1)
    new_text: str = Field(..., description="Nouveau texte qui remplacera l ancien")


@server.tool(name="edit_file")
async def edit_file(params: EditFileInput) -> str:
    """Modifie un fichier en remplacant un texte existant par un nouveau texte."""
    if ".." in params.filename or "/" in params.filename:
        return "Erreur : Le nom de fichier ne peut pas contenir .. ou /."

    filepath = FILES_DIR / params.filename
    if not filepath.exists():
        available = [f.name for f in FILES_DIR.iterdir() if f.is_file()]
        return f"Erreur : Fichier '{params.filename}' introuvable. Disponibles : {', '.join(available)}"

    content = filepath.read_text(encoding="utf-8")

    if params.old_text not in content:
        return f"Erreur : Le texte a remplacer n'a pas ete trouve dans '{params.filename}'. Utilisez read_file pour verifier le contenu exact."

    count = content.count(params.old_text)
    if count > 1:
        return f"Erreur : Le texte a remplacer apparait {count} fois. Fournissez un extrait plus precis pour eviter toute ambiguite."

    new_content = content.replace(params.old_text, params.new_text, 1)
    filepath.write_text(new_content, encoding="utf-8")

    return f"Fichier '{params.filename}' modifie avec succes.\n  Avant : {params.old_text}\n  Apres : {params.new_text}"


# Tool 6 : Lire et filtrer un CSV
class ReadCsvInput(BaseModel):
    filename: str = Field(..., description="Nom du fichier CSV (ex: equipe.csv)", min_length=1)
    filter_column: Optional[str] = Field(default=None, description="Nom de colonne pour filtrer (ex: statut, projet, nom)")
    filter_value: Optional[str] = Field(default=None, description="Valeur a chercher dans la colonne (ex: en cours, Marie)")
    columns: Optional[List[str]] = Field(default=None, description="Liste de colonnes a afficher (ex: ['nom', 'projet']). Toutes si vide.")
    limit: Optional[int] = Field(default=None, description="Nombre max de lignes a retourner", ge=1, le=1000)


@server.tool(name="read_csv")
async def read_csv_tool(params: ReadCsvInput) -> str:
    """Lit un fichier CSV avec filtrage optionnel par colonne et valeur.

    Permet de lire tout le CSV ou de filtrer par colonne.
    Exemples d appels :
      - Tout lire : filename=equipe.csv
      - Filtrer : filename=equipe.csv, filter_column=statut, filter_value=en cours
      - Colonnes specifiques : filename=equipe.csv, columns=[nom, projet, heures]
    """
    if ".." in params.filename or "/" in params.filename:
        return "Erreur : Le nom de fichier ne peut pas contenir .. ou /."

    filepath = FILES_DIR / params.filename
    if not filepath.exists():
        csv_files = [f.name for f in FILES_DIR.iterdir() if f.suffix == ".csv"]
        if csv_files:
            return f"Erreur : Fichier '{params.filename}' introuvable. CSV disponibles : {', '.join(csv_files)}"
        return f"Erreur : Fichier '{params.filename}' introuvable. Aucun CSV dans le repertoire."

    if not params.filename.endswith(".csv"):
        return "Erreur : Ce tool ne lit que les fichiers .csv. Utilisez read_file pour les autres formats."

    rows = []
    with open(filepath, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []

        if params.filter_column and params.filter_column not in headers:
            return f"Erreur : Colonne '{params.filter_column}' introuvable. Colonnes disponibles : {', '.join(headers)}"

        for row in reader:
            if params.filter_column and params.filter_value:
                cell = row.get(params.filter_column, "")
                if params.filter_value.lower() not in cell.lower():
                    continue
            rows.append(row)

    if params.limit:
        rows = rows[:params.limit]

    display_cols = params.columns if params.columns else headers
    invalid_cols = [c for c in display_cols if c not in headers]
    if invalid_cols:
        return f"Erreur : Colonnes introuvables : {', '.join(invalid_cols)}. Disponibles : {', '.join(headers)}"

    if not rows:
        return f"Aucun resultat dans '{params.filename}' avec ce filtre."

    lines = []
    lines.append(f"{len(rows)} ligne(s) dans '{params.filename}' :")
    lines.append("")
    lines.append(" | ".join(display_cols))
    lines.append("-" * (len(" | ".join(display_cols))))

    for row in rows:
        values = [str(row.get(c, "")) for c in display_cols]
        lines.append(" | ".join(values))

    lines.append("")
    lines.append(f"Colonnes disponibles : {', '.join(headers)}")

    return "\n".join(lines)


# Tool 7 : Modifier une cellule dans un fichier Excel
class EditXlsxInput(BaseModel):
    filename: str = Field(..., description="Nom du fichier Excel (ex: equipe.xlsx)", min_length=1)
    cell: str = Field(..., description="Reference de la cellule a modifier (ex: B3, E7, F10)", min_length=2, max_length=10)
    value: str = Field(..., description="Nouvelle valeur a ecrire dans la cellule")
    sheet: Optional[str] = Field(default=None, description="Nom de la feuille (premiere feuille si vide)")


@server.tool(name="edit_xlsx")
async def edit_xlsx(params: EditXlsxInput) -> str:
    """Modifie la valeur d une cellule dans un fichier Excel (.xlsx).

    Exemples :
      - Changer un nom : cell=B3, value=Jean Dupont
      - Changer un statut : cell=E7, value=termine
      - Changer un nombre : cell=F5, value=150
    """
    if ".." in params.filename or "/" in params.filename:
        return "Erreur : Le nom de fichier ne peut pas contenir .. ou /."

    filepath = FILES_DIR / params.filename
    if not filepath.exists():
        xlsx_files = [f.name for f in FILES_DIR.iterdir() if f.suffix == ".xlsx"]
        if xlsx_files:
            return f"Erreur : Fichier '{params.filename}' introuvable. Excel disponibles : {', '.join(xlsx_files)}"
        return f"Erreur : Fichier '{params.filename}' introuvable."

    if not params.filename.endswith(".xlsx"):
        return "Erreur : Ce tool ne modifie que les fichiers .xlsx."

    try:
        wb = load_workbook(str(filepath))
    except Exception as e:
        return f"Erreur : Impossible d'ouvrir '{params.filename}' : {e}"

    if params.sheet:
        if params.sheet not in wb.sheetnames:
            return f"Erreur : Feuille '{params.sheet}' introuvable. Feuilles disponibles : {', '.join(wb.sheetnames)}"
        ws = wb[params.sheet]
    else:
        ws = wb.active

    old_value = ws[params.cell].value

    # Convertir en nombre si possible
    new_value = params.value
    try:
        if "." in new_value:
            new_value = float(new_value)
        else:
            new_value = int(new_value)
    except ValueError:
        pass

    ws[params.cell] = new_value
    wb.save(str(filepath))

    return (
        f"Cellule {params.cell} modifiee dans '{params.filename}' :\n"
        f"  Avant : {old_value}\n"
        f"  Apres : {new_value}"
    )


# Resource : acces direct au contenu
@server.resource("file://{filename}")
async def get_file_resource(filename: str) -> str:
    """Expose le contenu d un fichier comme ressource MCP."""
    filepath = FILES_DIR / filename
    if not filepath.exists():
        return f"Fichier '{filename}' introuvable."
    return filepath.read_text(encoding="utf-8")


# Point d entree
if __name__ == "__main__":
    import sys
    print("Demarrage du serveur MCP 'files_mcp'...", file=sys.stderr)
    print(f"Repertoire surveille : {FILES_DIR}", file=sys.stderr)
    print(f"Tools : list_files, read_file, search_in_files, file_stats, edit_file, read_csv, edit_xlsx", file=sys.stderr)
    print(f"Resources : file://{{filename}}", file=sys.stderr)
    print(file=sys.stderr)
    server.run()
